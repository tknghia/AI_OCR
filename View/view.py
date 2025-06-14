import collections
import os
import sys
import base64
import io
from flask import Flask, render_template, request, jsonify, send_file, session, url_for, redirect
from docx import Document
from matplotlib import pyplot as plt
import threading
import time
from queue import Queue
import nbformat
from nbconvert import PythonExporter
from nbconvert.preprocessors import ExecutePreprocessor
import pandas as pd
from openpyxl.styles import PatternFill

# Add project root to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.append(project_root)

# Import controllers
from Controller.controller import ImageController, FileController, AuthController, MongoController

app = Flask(__name__)
app.secret_key = "thisismysecretkeyforthisapp"

# Initialize controllers
image_controller = ImageController()
file_controller = FileController()
auth_controller = AuthController()
mongo_controller = MongoController()


# Training queue and thread
training_queue = Queue()
training_thread = None
def train_model():
    while True:
        # Wait for a training task
        task = training_queue.get()
        if task is None:
            break
        
        print("Starting model training...")
        # Ở đây, bạn sẽ chạy mã training từ Jupyter notebook của bạn
        
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(current_dir)
        notebook_path = os.path.join(project_root, 'idvn-ocr.ipynb')
        # Đọc nội dung notebook
        with open(notebook_path, 'r', encoding='utf-8') as file:
            notebook_content = nbformat.read(file, as_version=4)

        # Sử dụng ExecutePreprocessor để thực thi notebook
        ep = ExecutePreprocessor(timeout=600, kernel_name='python3')
        ep.preprocess(notebook_content, {'metadata': {'path': project_root}})

        time.sleep(10)  # Giả lập thời gian training
        print("Model training completed")
        
        training_queue.task_done()

def start_training_thread():
    global training_thread
    if training_thread is None or not training_thread.is_alive():
        training_thread = threading.Thread(target=train_model)
        training_thread.start()

@app.route('/')
def index():
    if session.get('logged_in'):
        is_admin=session['is_admin']
        return render_template('index.html',is_admin=is_admin)
    return redirect(url_for('login'))

@app.route('/export_excel')
def export_excel():
    users = mongo_controller.get_all_users()
    return render_template('export_excel.html', users=users)

@app.route('/export_excel1')
def export_excel1():
    users = mongo_controller.get_all_users() 

    # Chuyển dữ liệu người dùng hiện có thành DataFrame của pandas
    data = [
        {
            "Username": user.get("username", ""),  
            "Email": user.get("email", ""), 
            "ID_Card": user.get("id_card", ""),  
            "Tên": user.get("name", ""), 
            "Địa Chỉ": user.get("current_place", ""),  
            "Ngày Sinh": user.get("dob", ""),  
            "Quốc Tịch": user.get("nationality", ""),  
            "Trạng Thái KYC": 'Đã xác minh' if user.get("is_kyc") else 'Chưa xác minh',
        }
        for user in users
    ]

    df = pd.DataFrame(data)

    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Users')
        workbook = writer.book
        worksheet = workbook['Users']

        # Format Header Row
        from openpyxl.styles import Alignment, Font, Border, Side
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col in worksheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(df.columns)):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # Format Data Rows
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # Apply Conditional Formatting for KYC Column
        kyc_column_index = 8  # "Trạng Thái KYC" column
        for row in range(2, len(df) + 2): 
            kyc_cell = worksheet.cell(row=row, column=kyc_column_index)
            if kyc_cell.value == 'Đã xác minh':
                kyc_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  
            else:
                kyc_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  

        # Auto-adjust Column Width
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter  # Get column letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return send_file(output, as_attachment=True, download_name="users.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



@app.route('/result')
def result():
    # Kiểm tra nếu người dùng đã đăng nhập
    user_id = session.get('user_id')
    if 'logged_in' in session and session['logged_in']:
        # Lấy thông tin người dùng từ MongoDB qua AuthController
        user_info = auth_controller.get_user_info(user_id)
        return render_template('result.html', user=user_info)
    else:
        return render_template('login.html')

@app.route('/convert', methods=['POST'])
def handle_convert_images():
    document_type = request.form.get('documentType')
    update_profile = request.form.get('KYC_PROFILE')  # Lấy giá trị documentType
    files = request.files.getlist('images')
    user_id = session.get('user_id')
    return image_controller.process_images(files,document_type,update_profile, userId=user_id)


@app.route('/download', methods=['POST'])
def download():
    content = request.json.get('content', '')
    return file_controller.download_content(content)

@app.route('/save', methods=['POST'])
def save():
    labels = request.json.get('labels', [])
    result_id = request.json.get('result_id', '')
    user_id=session['user_id']
    result = file_controller.save_labels(labels, result_id,user_id)
    # Thêm task training vào queue
    # training_queue.put(True)
    # Đảm bảo thread training đang chạy
    # start_training_thread()
    return result

def read_word_file(file_path):
    doc = Document(file_path)
    return [[cell.text for cell in row.cells] for table in doc.tables for row in table.rows]

@app.route('/report_log')
def report_log():
    # Check if the user is logged in and has a user_id in the session
    user_id = session.get('user_id')
    if user_id:
        # Retrieve logs made by this user
        logs = mongo_controller.get_predictions_by_user_id(user_id)
        
        if isinstance(logs, str):  # If the result is an error message
            return logs, 404
        
        # Process logs into a format suitable for rendering
        processed_logs = []
        for log in logs:
            processed_log = {
                "upload_time": log.get("upload_time"),
                "average_accuracy": log.get("average_accuracy"),
                "samples": [
                    {
                        "index": idx + 1,
                        "label": image.get("label", ""),
                        "prediction": image.get("prediction", ""),
                        "accuracy": image.get("accuracy", "")
                    }
                    for idx, image in enumerate(log.get("list_images", []))
                ]
            }
            processed_logs.append(processed_log)
        is_admin=session['is_admin']
        return render_template('report_log.html', logs=processed_logs,is_admin=is_admin)

    return "User not logged in or no user_id found in session.", 403



@app.route('/chart')
def view_chart():
    # Retrieve the user ID from the session
    user_id = session.get('user_id')
    if not user_id:
        return "User not logged in.", 403

    # Get all logs for the user
    logs = mongo_controller.get_predictions_by_user_id(user_id)
    if isinstance(logs, str):
        return logs, 404  # If an error message was returned

    # Calculate average accuracy and loss for each log entry
    samples = []
    average_accuracies = []
    loss_percentages = []

    for log_index, log in enumerate(logs, start=1):
        total_accuracy = 0
        total_samples = len(log.get("list_images", []))

        for image in log.get("list_images", []):
            accuracy = image.get("accuracy", 0)
            total_accuracy += accuracy

        if total_samples > 0:
            average_accuracy = total_accuracy / total_samples
            loss_percentage = 100 - average_accuracy
        else:
            average_accuracy = 0
            loss_percentage = 0

        # Append data for chart
        samples.append(f'#{log_index}')
        average_accuracies.append(average_accuracy)
        loss_percentages.append(loss_percentage)

    # Generate the plot
    plt.figure(figsize=(12, 6))
    plt.plot(samples, average_accuracies, marker='o', label='Average Accuracy (%)', color='blue')
    plt.plot(samples, loss_percentages, marker='o', label='Loss Percentage (%)', color='red')
    plt.xlabel('Samples')
    plt.ylabel('Percentage (%)')
    plt.title('Average Accuracy and Loss Percentage by Log Entry')
    plt.legend()
    plt.grid(True)

    # Convert plot to PNG image
    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    buf.seek(0)
    graph_url = base64.b64encode(buf.getvalue()).decode('utf-8')
    buf.close()
    is_admin=session['is_admin']
    # Pass the chart to the HTML template
    return render_template('chart.html', graph_url=graph_url,is_admin=is_admin)


@app.route('/register', methods=["POST"])
def register():
    data = request.get_json()
    result = auth_controller.register(data)
    return jsonify(result)

@app.route('/login', methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json() if request.is_json else request.form
        result = auth_controller.login(data.get("username"), data.get("password"))
        # Check for a successful login response with user_id
        if result.get("message") == "Login successful.":
            session['logged_in'] = True
            session['username'] = data.get("username")
            session['is_admin'] = result.get("is_admin")  # Ensure this is being set
            session['user_id'] = result.get("user_id")  # Store user ID in the session
            return jsonify({"message": result["message"], "user_id": result["user_id"]})
        
        # Return error response if login failed
        return jsonify({"error": result.get("error")}), 401
    return render_template("login.html")

@app.route('/summary')
def summary_of_model():
    # Retrieve the user ID from the session
    user_id = session.get('user_id')
    if not user_id:
        return "User not logged in.", 403

    # Get all logs for the user
    logs = mongo_controller.get_predictions_by_user_id(user_id)
    if isinstance(logs, str):
        return logs, 404  # If an error message was returned

    # Calculate average accuracy and loss
    total_accuracy = 0
    total_samples = 0
    failed_samples = []
    correct_samples = []

    # Loop through logs with entry number
    for log_index, log in enumerate(logs, start=1):
        for image in log.get("list_images", []):
            accuracy = image.get("accuracy", 0)
            total_accuracy += accuracy
            total_samples += 1  # Increment total_samples for each sample processed
            
            sample_info = {
                "log_entry": log_index,  # Add log entry number
                "label": image.get("label"),
                "prediction": image.get("prediction"),
                "accuracy": accuracy
            }
            
            # Separate correct and failed samples
            if accuracy < 100:
                failed_samples.append(sample_info)
            else:
                correct_samples.append(sample_info)

    # Calculate average accuracy and loss percentage
    if total_samples > 0:
        average_accuracy = total_accuracy / total_samples
        loss_percentage = 100-average_accuracy
    else:
        average_accuracy = 0
        loss_percentage = 0
    is_admin=session['is_admin']
    # Pass the calculated data and samples to the template
    return render_template('summary.html', total_samples=len(logs), 
                           average_accuracy=average_accuracy, 
                           loss_percentage=loss_percentage, 
                           failed_samples=failed_samples,
                           correct_samples=correct_samples,
                           is_admin=is_admin
                           )


@app.route("/kyc", methods=["POST"])
def KycInfo():
    user_id = session.get('user_id')
    if not user_id:
        return {"error": "User not logged in"}, 403

    infos = request.json.get('info', [])
    result = mongo_controller.update_kyc_info(user_id, infos)
    return result

def start_app():
    #start_training_thread()
    app.run(host="0.0.0.0", port=5001, debug=False)
if __name__ == '__main__':
    start_app()
